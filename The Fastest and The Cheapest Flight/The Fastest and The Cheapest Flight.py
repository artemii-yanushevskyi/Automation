import openpyxl

class Flight:
    def __init__(self, info):
        print(info)
        number, cityDep, cityArr, timeDep, timeArr, price = info
        self.number = number
        self.cityDep = cityDep
        self.cityArr = cityArr
        self.timeDep = timeDep
        self.timeArr = timeArr
        self.price = price
        
    def __str__(self):
        ''' str or print '''
        return self.number
        
    def __repr__(self):
        ''' repr in console'''
        return str(self.number)+' '+str(self.price)


class Journey:
    def __init__(self, route):
        self.route = route
        
    def price(self):
        return sum([f.price for f in self.route])
    
    def time(self):
        return self.route[-1].timeArr - self.route[0].timeDep

    def __str__(self):
        return str(self.route)

wb = openpyxl.load_workbook('Book.xlsx')
print(wb.get_sheet_names())
ws = wb['Sheet1']
flights = []
for i in range(ws.min_row + 1, ws.max_row + 1):
    flight = Flight([ws.cell(row = i, column = j).value for j in range(ws.min_column, ws.max_column + 1)])
    print(flight, end=' ')
    flights.append(flight)
print()
print()
f = 'K'
t = 'US'
time = 0
journeys = []

def findJourneys(flights, f, t, time, history):
    appropriateFlights = [flight for flight in flights if time<flight.timeDep]
    for flight in appropriateFlights:
        if f == flight.cityDep and flight.cityArr == t:
            h = history + [flight]
            journeys.append(Journey(h))
        elif f == flight.cityDep:
            h = history + [flight]
            findJourneys(appropriateFlights, flight.cityArr, t, flight.timeArr, h)
        else:
            pass

findJourneys(flights, f, t, time, [])
for j in journeys:
    print('The route', j ,'will cost $' + str(j.price()))
cheapestWay = min([j.price() for j in journeys])
fastestWay = min([j.time() for j in journeys])
print('The cheapest way will be $' + str(cheapestWay))
print('The fastest way will be' ,fastestWay, 'of something')
